# Create your views here.
from rest_framework import viewsets, status
from rest_framework.parsers import MultiPartParser, FileUploadParser, FormParser, JSONParser
from rest_framework.response import Response
from rest_framework_jwt.settings import api_settings
from rest_framework_jwt.utils import jwt_decode_handler
from backoffice.utils.mailing import sendmail_to_customer, RESET_PASSWORD_MAIL, VALIDATION_ACCOUNT
from backoffice.models import Manager
from backoffice.utils.constant import FRONTEND_URL, ACCOUNT_ACTIVATION_URL, PASSWORD_RESET_URL

from backoffice.models import *
from backoffice.serializers import *
from backoffice.utils.function import *
from backoffice.utils.function import make_video

jwt_payload_handler = api_settings.JWT_PAYLOAD_HANDLER
jwt_encode_handler = api_settings.JWT_ENCODE_HANDLER

from rest_framework.decorators import action, api_view, permission_classes
from rest_framework import permissions
from openpyxl import load_workbook
from django.contrib.auth.hashers import check_password, make_password
from django.contrib.auth.models import User
from backoffice.serializers import *
from backoffice.utils.mapper import *
from backoffice.models import *
from backoffice.utils.function import *
from backoffice.utils.mapper import NUMBER

"""
API endpoint that allows users to be viewed or edited
"""


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all().order_by('-date_joined')
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAdminUser]

    # Function allow user to change password
    @action(methods=['post'], detail=True, url_name='change_password')
    def change_password(self, request, pk=None):
        try:
            data = request.data
            old_password = data['old_password']
            new_password = data['new_password']
            user = User.objects.get(id=pk)
            if not check_password(old_password, user.password):
                return Response({'old_password': ["Wrong password"]}, status=status.HTTP_400_BAD_REQUEST)
            password = make_password(new_password)
            user.password = password
            user.save()
            return Response({"password change successfully"}, status=status.HTTP_200_OK)
        except:
            return Response(status=status.HTTP_404_NOT_FOUND)


class AttributeViewSet(viewsets.ModelViewSet):
    queryset = Attribute.objects.all().order_by('name')
    serializer_class = AttributeSerializer
    permission_classes = [permissions.IsAdminUser]


class CountryViewSet(viewsets.ModelViewSet):
    queryset = Country.objects.all().order_by('name')
    serializer_class = CountrySerializer
    permission_classes = [permissions.IsAuthenticated]


class ManagerViewSet(viewsets.ModelViewSet):
    queryset = Manager.objects.all()
    serializer_class = ManagerSerializer
    permission_classes = [permissions.IsAuthenticated]


class StaffViewSet(viewsets.ModelViewSet):
    queryset = Staff.objects.all()
    serializer_class = StaffSerializer
    permission_classes = [permissions.IsAuthenticated]


class PlayerViewSet(viewsets.ModelViewSet):
    queryset = Player.objects.all().order_by('number')
    serializer_class = PlayerSerializer
    permission_classes = [permissions.IsAuthenticated]


class ActionViewSet(viewsets.ModelViewSet):
    queryset = Action.objects.all().order_by('name')
    serializer_class = ActionSerializer
    permission_classes = [permissions.IsAuthenticated]


class TeamViewSet(viewsets.ModelViewSet):
    queryset = Team.objects.all().order_by('name')
    serializer_class = TeamSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(methods=['get'], detail=True, url_name='players')
    def players(self, request, pk=None):
        try:
            players = Player.objects.filter(current_team_id=pk)
            players = PlayerSerializer(players, many=True, context={'request': request})
            players = players.data
            return Response(players, status=status.HTTP_200_OK)
        except:
            return Response({"error": "error occur"}, status=status.HTTP_404_NOT_FOUND)


class GameViewSet(viewsets.ModelViewSet):
    queryset = Game.objects.all().order_by('date')
    serializer_class = GameSerializer
    permission_classes = [permissions.IsAdminUser]

    @action(methods=['get'], detail=True, url_name='players_actions')
    def players_actions(self, request, pk=None):

        try:
            players_actions = PlayerActionInGame.objects.filter(game_id=pk)
            players_actions = PlayerActionInGameSerializer(players_actions, many=True, context={'request': request})
            players_actions = players_actions.data
            return Response(players_actions, status=status.HTTP_200_OK)
        except:
            return Response({'error': 'error occur'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GameVideoViewSet(viewsets.ModelViewSet):
    parser_classes = (MultiPartParser, FormParser,)
    queryset = GameVideo.objects.all().order_by('id')
    serializer_class = GameVideoSerializer
    permission_classes = [permissions.IsAdminUser]


class PeriodViewSet(viewsets.ModelViewSet):
    queryset = Period.objects.all()
    serializer_class = PeriodSerializer
    permission_classes = [permissions.IsAuthenticated]


class SeasonViewSet(viewsets.ModelViewSet):
    queryset = Season.objects.all().order_by('begin')
    serializer_class = SeasonSerializer
    permission_classes = [permissions.IsAuthenticated]


class LeagueViewSet(viewsets.ModelViewSet):
    queryset = League.objects.all().order_by('name')
    serializer_class = LeagueSerializer
    permission_classes = [permissions.IsAuthenticated]


class PlayerActionInGameViewSet(viewsets.ModelViewSet):
    queryset = PlayerActionInGame.objects.all().order_by('video_start')
    serializer_class = PlayerActionInGameSerializer
    permission_classes = [permissions.IsAdminUser]


class ContractViewSet(viewsets.ModelViewSet):
    queryset = Contract.objects.all().order_by('team')
    serializer_class = ContractSerializer
    permission_classes = [permissions.IsAdminUser]


class PlayerPositionViewSet(viewsets.ModelViewSet):
    queryset = PlayerPosition.objects.all().order_by('code')
    serializer_class = PlayerPositionSerializer
    permission_classes = [permissions.IsAdminUser]


class TeamExcelViewSet(viewsets.ViewSet):
    serializer_class = TeamExcelSerializer

    def create(self, request, pk=None):
        file = request.FILES.get('file')
        sht = load_workbook(file)
        sheet = sht.active
        global team
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            team = team_maker(row)
        for row in sheet.iter_rows(min_row=4, values_only=True):
            country = country_maker(row)
            if type(row[NUMBER]) == int:
                player = player_maker(row, country)
                period_maker(row, player, team)
            else:
                staff = staff_maker(row, country)
                contract_maker(row, staff, team)
        response = "POST API and you have uploaded a {} file"
        return Response(response)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def get_password_reset_token(request):
    data = request.data['email']
    try:
        user = User.objects.get(email=data)
        payload = jwt_payload_handler(user)
        token = jwt_encode_handler(payload)
        sendmail_to_customer(token, user, RESET_PASSWORD_MAIL, FRONTEND_URL, PASSWORD_RESET_URL)
        content = {
            "message": "un mail de recuperation de compte a ete envoye a l'adresse indique!"
        }
        return Response(content)
    except:
        return Response({"message": "user with this email not found"}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def register(request):
    new_manager = Manager()
    new_user = User()
    data = request.data
    new_manager.last_name = data['lastName']
    new_manager.first_name = data['firstName']
    new_manager.age = data['age']
    new_manager.status = data['status']
    new_user.username = data['username']
    new_user.email = data['email']
    new_user.password = make_password(data['password'])
    new_user.is_active = False
    new_manager.account = new_user

    try:
        new_user.save()
        new_manager.save()
        payload = jwt_payload_handler(new_user)
        token = jwt_encode_handler(payload)
        sendmail_to_customer(token, new_user, VALIDATION_ACCOUNT, FRONTEND_URL, ACCOUNT_ACTIVATION_URL)
        return Response(
            {
                "message": "un mail d'activation a ete envoyer sur le mail de creation de votre compte"
            },
            status=status.HTTP_200_OK
        )
    except Exception as exception:
        print(exception)
        return Response({"error": repr(exception)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PATCH'])
@permission_classes([permissions.AllowAny])
def active_user_account(request):
    token = request.data['token']
    try:
        identity = jwt_decode_handler(token)
        print(identity)
        user = User.objects.get(pk=identity['user_id'])
        user.is_active = True
        user.save()
        return Response({'message': 'compte active avec success'}, status=status.HTTP_200_OK)
    except Exception as exception:
        return Response({"error": repr(exception)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def reset_password(request):
    token = request.data['token']
    password = request.data['password']
    try:
        identity = jwt_decode_handler(token)
        user = User.objects.get(email=identity['email'])
        hash_pwd = make_password(password)
        user.password = hash_pwd
        user.save()
        return Response({"message": "mot de passe configurer avec success"}, status=status.HTTP_200_OK)
    except Exception as exception:
        return Response({"error": repr(exception)}, status=status.HTTP_404_NOT_FOUND)
